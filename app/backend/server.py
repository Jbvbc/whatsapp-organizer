from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials, APIKeyHeader
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
try:
    from motor.motor_asyncio import AsyncIOMotorClient
    MOTOR_AVAILABLE = True
except ImportError:
    AsyncIOMotorClient = None
    MOTOR_AVAILABLE = False

from mock_db import get_mock_db
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from bson import ObjectId
import json
import zipfile
from io import BytesIO
from datetime import datetime, timezone, timedelta
from jose import jwt, JWTError
from passlib.context import CryptContext
import asyncio
import secrets
import hashlib
import time
from collections import defaultdict, deque

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Database connection — MongoDB ou MockDB (JSON local)
if os.environ.get('DB_TYPE') == 'mock' or not MOTOR_AVAILABLE:
    logging.warning("Usando MockDB (armazenamento local em data.json)")
    db = get_mock_db()
else:
    mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
    client = AsyncIOMotorClient(mongo_url)
    db = client[os.environ.get('DB_NAME', 'whatsapp_organizer')]
    logging.info("Usando MongoDB")

# Create the main app
app = FastAPI(
    title="WhatsApp Contact Organizer API",
    description="REST API for managing contacts, groups, events, and scheduled messages with multi-organization support, JWT authentication, and API key integration.",
    version="1.0.0",
    contact={
        "name": "API Support",
        "url": "https://github.com/Jbvbc/whatsapp-organizer",
    },
    openapi_tags=[
        {"name": "Auth", "description": "User registration, login, and profile"},
        {"name": "Contacts", "description": "Manage contacts (CRUD, sync, search)"},
        {"name": "Groups", "description": "Manage contact groups"},
        {"name": "Events", "description": "Manage events and birthday notifications"},
        {"name": "Scheduled Messages", "description": "Schedule and manage WhatsApp messages"},
        {"name": "Organizations", "description": "Multi-organization management (admin only)"},
        {"name": "API Keys", "description": "API key management for integrations (admin only)"},
        {"name": "Reports", "description": "Dashboard statistics and activity reports"},
        {"name": "Backup & Export", "description": "Backup, restore, and data export"},
        {"name": "CRM", "description": "CRM integration (HubSpot, Salesforce) for contact sync"},
        {"name": "Webhooks", "description": "Webhook configuration for event-driven integrations"},
        {"name": "WhatsApp", "description": "WhatsApp Business API integration for message sending and delivery tracking"},
    ],
)
api_router = APIRouter(prefix="/api")

# ─── Rate Limiter ───────────────────────────────────────
class RateLimiter:
    def __init__(self, requests_per_minute: int = 60):
        self.rpm = requests_per_minute
        self._buckets: dict[str, deque] = defaultdict(lambda: deque())

    def check(self, key: str) -> bool:
        now = time.time()
        bucket = self._buckets[key]
        # Purge old entries
        while bucket and bucket[0] < now - 60:
            bucket.popleft()
        if len(bucket) >= self.rpm:
            return False
        bucket.append(now)
        return True

rate_limiter = RateLimiter(120)  # 120 requests/minute per client

class RateLimitMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        # Skip rate limiting for docs and openapi
        path = request.url.path
        if path in ("/docs", "/openapi.json", "/redoc"):
            return await call_next(request)

        client_key = request.client.host if request.client else "unknown"
        # Use API key or JWT user for more granular limiting
        api_key = request.headers.get("x-api-key")
        if api_key:
            client_key = f"apikey:{api_key[:8]}"
        elif "authorization" in request.headers:
            client_key = f"jwt:{client_key}"

        if not rate_limiter.check(client_key):
            raise HTTPException(status_code=429, detail="Too many requests. Please try again later.")
        return await call_next(request)

# ─── API Key Helpers ────────────────────────────────────
api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)

def hash_api_key(key: str) -> str:
    return hashlib.sha256(key.encode()).hexdigest()

def generate_api_key() -> str:
    return f"wco_{secrets.token_urlsafe(32)}"

async def get_api_key_user(api_key: str = Depends(api_key_header)):
    """Dependency to authenticate via X-API-Key header (alternative to JWT)"""
    if not api_key:
        return None
    hashed = hash_api_key(api_key)
    doc = await db.api_keys.find_one({"keyHash": hashed, "isActive": True})
    if not doc:
        raise HTTPException(status_code=401, detail="Invalid API key")
    # Update last used
    await db.api_keys.update_one({"_id": doc["_id"]}, {"$set": {"lastUsedAt": datetime.utcnow()}})
    return {
        "id": str(doc["_id"]),
        "name": doc.get("name", "API Key"),
        "role": "editor",  # API keys get editor-level access
        "scopes": doc.get("scopes", []),
        "source": "api_key"
    }

# Auth config
SECRET_KEY = os.environ.get("JWT_SECRET", "super-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 30
pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
security = HTTPBearer()

class UserCreate(BaseModel):
    email: str
    password: str
    name: str

class UserLogin(BaseModel):
    email: str
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """FastAPI dependency that returns current user from JWT token"""
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return {
        "id": str(user["_id"]),
        "email": user["email"],
        "name": user["name"],
        "role": user.get("role", "viewer")
    }

def require_role(*roles: str):
    """Dependency factory: require specific roles"""
    async def role_checker(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

async def get_current_user_or_api_key(
    jwt_user: dict = Depends(get_current_user),
    api_key_user: dict = Depends(get_api_key_user),
):
    """Accept either JWT or API key authentication"""
    if api_key_user:
        return api_key_user
    return jwt_user

# Auth Routes
@api_router.post("/auth/register", response_model=TokenResponse, tags=["Auth"], summary="Register a new user")
async def register(user: UserCreate):
    """Register a new user. First user becomes admin, subsequent users become viewer."""
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_dict = {
        "email": user.email,
        "name": user.name,
        "password": get_password_hash(user.password),
        "role": "admin",  # First user is admin
        "createdAt": datetime.utcnow(),
        "updatedAt": datetime.utcnow()
    }
    
    # Check if this is the first user, otherwise new users get "viewer" role
    user_count = await db.users.count_documents({})
    if user_count > 0:
        user_dict["role"] = "viewer"
    
    result = await db.users.insert_one(user_dict)
    user_dict["_id"] = result.inserted_id
    
    access_token = create_access_token({"sub": str(result.inserted_id), "role": user_dict["role"]})
    return TokenResponse(
        access_token=access_token,
        user=UserResponse(
            id=str(result.inserted_id),
            email=user_dict["email"],
            name=user_dict["name"],
            role=user_dict["role"]
        )
    )

@api_router.post("/auth/login", response_model=TokenResponse, tags=["Auth"], summary="Login and get JWT token")
async def login(user: UserLogin):
    """Authenticate with email and password. Returns a JWT token valid for 30 days."""
    db_user = await db.users.find_one({"email": user.email})
    if not db_user or not verify_password(user.password, db_user["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    access_token = create_access_token({
        "sub": str(db_user["_id"]),
        "role": db_user.get("role", "viewer")
    })
    return TokenResponse(
        access_token=access_token,
        user=UserResponse(
            id=str(db_user["_id"]),
            email=db_user["email"],
            name=db_user["name"],
            role=db_user.get("role", "viewer")
        )
    )

@api_router.get("/auth/me", response_model=UserResponse, tags=["Auth"], summary="Get current user profile")
async def get_me(user: dict = Depends(get_current_user)):
    """Returns the authenticated user's profile information including email, name, and role."""
    return UserResponse(**user)

# Models
class Contact(BaseModel):
    id: Optional[str] = None
    name: str
    phone: str
    photo: Optional[str] = None  # base64
    notes: Optional[str] = ""
    tags: List[str] = []
    isFavorite: bool = False
    rawContactId: Optional[str] = None
    organizationId: Optional[str] = None
    createdAt: Optional[datetime] = None
    updatedAt: Optional[datetime] = None

class ContactUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    photo: Optional[str] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = None
    isFavorite: Optional[bool] = None
    birthday: Optional[str] = None  # ISO format date

class Event(BaseModel):
    id: Optional[str] = None
    title: str
    description: Optional[str] = None
    date: datetime
    type: str  # birthday, anniversary, custom
    contactId: Optional[str] = None
    organizationId: Optional[str] = None
    isRecurring: bool = False
    recurringPattern: Optional[str] = None
    isActive: bool = True
    createdAt: Optional[datetime] = None
    updatedAt: Optional[datetime] = None

class EventUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    date: Optional[datetime] = None
    type: Optional[str] = None
    contactId: Optional[str] = None
    isRecurring: Optional[bool] = None
    recurringPattern: Optional[str] = None
    isActive: Optional[bool] = None

class Group(BaseModel):
    id: Optional[str] = None
    name: str
    color: Optional[str] = "#4A90E2"
    contactIds: List[str] = []
    organizationId: Optional[str] = None
    createdAt: Optional[datetime] = None

class GroupUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    contactIds: Optional[List[str]] = None

class ScheduledMessage(BaseModel):
    id: Optional[str] = None
    groupId: str
    message: str
    scheduledTime: datetime
    isRecurring: bool = False
    recurringPattern: Optional[str] = None  # daily, weekly, monthly
    isActive: bool = True
    status: str = "pending"  # pending, sent, failed
    organizationId: Optional[str] = None
    createdAt: Optional[datetime] = None
    updatedAt: Optional[datetime] = None

class ScheduledMessageUpdate(BaseModel):
    message: Optional[str] = None
    scheduledTime: Optional[datetime] = None
    isRecurring: Optional[bool] = None
    recurringPattern: Optional[str] = None
    isActive: Optional[bool] = None
    status: Optional[str] = None

class Organization(BaseModel):
    id: Optional[str] = None
    name: str
    color: Optional[str] = "#4A90E2"
    description: Optional[str] = ""
    createdAt: Optional[datetime] = None
    updatedAt: Optional[datetime] = None

class OrganizationUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    description: Optional[str] = None

class ApiKeyCreate(BaseModel):
    name: str = Field(description="Friendly name for this API key")
    scopes: List[str] = Field(default=[], description="Allowed scopes (e.g. ['contacts:read', 'messages:write'])")

class ApiKeyResponse(BaseModel):
    id: str
    name: str
    key: Optional[str] = None  # Only returned on creation
    scopes: List[str] = []
    isActive: bool = True
    lastUsedAt: Optional[datetime] = None
    createdAt: datetime

class CrmIntegrationCreate(BaseModel):
    provider: str = Field(description="CRM provider (hubspot, salesforce)")
    name: str = Field(description="Friendly name for this integration")
    apiKey: str = Field(description="API key/token for the CRM")
    apiUrl: Optional[str] = Field(None, description="Custom API URL if applicable")
    organizationId: Optional[str] = None

class CrmIntegrationUpdate(BaseModel):
    name: Optional[str] = None
    apiKey: Optional[str] = None
    apiUrl: Optional[str] = None
    isActive: Optional[bool] = None

class CrmIntegrationResponse(BaseModel):
    id: str
    provider: str
    name: str
    apiUrl: Optional[str] = None
    isActive: bool
    lastSyncAt: Optional[datetime] = None
    lastSyncStatus: Optional[str] = None
    organizationId: Optional[str] = None
    createdAt: datetime
    updatedAt: datetime

class CrmProviderInfo(BaseModel):
    id: str
    name: str
    description: str

class WebhookCreate(BaseModel):
    url: str = Field(description="URL to send webhook POST requests to")
    events: List[str] = Field(description="Events to subscribe to: contact.created, contact.updated, message.scheduled, message.sent, event.created, event.upcoming")
    name: Optional[str] = Field(None, description="Friendly name")
    secret: Optional[str] = Field(None, description="Secret for HMAC signature (optional)")

class WebhookUpdate(BaseModel):
    url: Optional[str] = None
    events: Optional[List[str]] = None
    name: Optional[str] = None
    secret: Optional[str] = None
    isActive: Optional[bool] = None

class WebhookResponse(BaseModel):
    id: str
    url: str
    events: List[str]
    name: Optional[str] = None
    isActive: bool
    lastTriggeredAt: Optional[datetime] = None
    lastResponseStatus: Optional[int] = None
    createdAt: datetime
    updatedAt: datetime

class WhatsAppConfigCreate(BaseModel):
    phoneNumberId: str = Field(description="WhatsApp Business phone number ID")
    accessToken: str = Field(description="WhatsApp Cloud API access token (long-lived)")
    businessAccountId: Optional[str] = Field(None, description="WhatsApp Business Account ID")
    webhookSecret: Optional[str] = Field(None, description="Verify token for WhatsApp webhook callback")
    organizationId: Optional[str] = None

class WhatsAppConfigUpdate(BaseModel):
    phoneNumberId: Optional[str] = None
    accessToken: Optional[str] = None
    businessAccountId: Optional[str] = None
    webhookSecret: Optional[str] = None
    isActive: Optional[bool] = None

class WhatsAppConfigResponse(BaseModel):
    id: str
    phoneNumberId: str
    businessAccountId: Optional[str] = None
    isActive: bool
    organizationId: Optional[str] = None
    createdAt: datetime
    updatedAt: datetime

class WhatsAppMessageStatus(BaseModel):
    id: str
    externalMessageId: str
    scheduledMessageId: Optional[str] = None
    recipientPhone: str
    status: str  # sent, delivered, read, failed
    timestamp: datetime

# Helper function to convert ObjectId to string
def contact_helper(contact) -> dict:
    return {
        "id": str(contact["_id"]),
        "name": contact["name"],
        "phone": contact["phone"],
        "photo": contact.get("photo"),
        "notes": contact.get("notes", ""),
        "tags": contact.get("tags", []),
        "isFavorite": contact.get("isFavorite", False),
        "rawContactId": contact.get("rawContactId"),
        "organizationId": contact.get("organizationId"),
        "createdAt": contact.get("createdAt"),
        "updatedAt": contact.get("updatedAt")
    }

def group_helper(group) -> dict:
    return {
        "id": str(group["_id"]),
        "name": group["name"],
        "color": group.get("color", "#4A90E2"),
        "contactIds": group.get("contactIds", []),
        "organizationId": group.get("organizationId"),
        "createdAt": group.get("createdAt")
    }

def scheduled_message_helper(scheduled_message) -> dict:
    return {
        "id": str(scheduled_message["_id"]),
        "groupId": scheduled_message["groupId"],
        "message": scheduled_message["message"],
        "scheduledTime": scheduled_message["scheduledTime"],
        "isRecurring": scheduled_message.get("isRecurring", False),
        "recurringPattern": scheduled_message.get("recurringPattern"),
        "isActive": scheduled_message.get("isActive", True),
        "status": scheduled_message.get("status", "pending"),
        "organizationId": scheduled_message.get("organizationId"),
        "createdAt": scheduled_message.get("createdAt"),
        "updatedAt": scheduled_message.get("updatedAt")
    }

def organization_helper(org) -> dict:
    return {
        "id": str(org["_id"]),
        "name": org["name"],
        "color": org.get("color", "#4A90E2"),
        "description": org.get("description", ""),
        "createdAt": org.get("createdAt"),
        "updatedAt": org.get("updatedAt")
    }

def api_key_helper(doc) -> dict:
    return {
        "id": str(doc["_id"]),
        "name": doc["name"],
        "scopes": doc.get("scopes", []),
        "isActive": doc.get("isActive", True),
        "lastUsedAt": doc.get("lastUsedAt"),
        "createdAt": doc["createdAt"]
    }

def crm_integration_helper(doc) -> dict:
    return {
        "id": str(doc["_id"]),
        "provider": doc["provider"],
        "name": doc["name"],
        "apiUrl": doc.get("apiUrl"),
        "isActive": doc.get("isActive", True),
        "lastSyncAt": doc.get("lastSyncAt"),
        "lastSyncStatus": doc.get("lastSyncStatus"),
        "organizationId": doc.get("organizationId"),
        "createdAt": doc["createdAt"],
        "updatedAt": doc["updatedAt"]
    }

def webhook_helper(doc) -> dict:
    return {
        "id": str(doc["_id"]),
        "url": doc["url"],
        "events": doc.get("events", []),
        "name": doc.get("name"),
        "isActive": doc.get("isActive", True),
        "lastTriggeredAt": doc.get("lastTriggeredAt"),
        "lastResponseStatus": doc.get("lastResponseStatus"),
        "createdAt": doc["createdAt"],
        "updatedAt": doc["updatedAt"]
    }

async def dispatch_webhook_event(event_type: str, payload: dict):
    """Fire webhooks subscribed to a given event type"""
    try:
        webhooks = await db.webhooks.find({
            "events": event_type,
            "isActive": True
        }).to_list(50)
    except Exception:
        return  # Collection might not exist yet
    
    for wh in webhooks:
        asyncio.create_task(send_webhook(wh, event_type, payload))

async def send_webhook(wh: dict, event_type: str, payload: dict):
    """Send a single webhook POST request"""
    import httpx
    url = wh["url"]
    body = json.dumps({"event": event_type, "data": payload, "timestamp": datetime.utcnow().isoformat()}, default=str)
    headers = {"Content-Type": "application/json"}
    secret = wh.get("secret")
    if secret:
        import hmac
        signature = hmac.new(secret.encode(), body.encode(), hashlib.sha256).hexdigest()
        headers["X-Webhook-Signature"] = signature
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(url, content=body, headers=headers)
        await db.webhooks.update_one(
            {"_id": wh["_id"]},
            {"$set": {"lastTriggeredAt": datetime.utcnow(), "lastResponseStatus": resp.status_code, "updatedAt": datetime.utcnow()}}
        )
    except Exception as e:
        await db.webhooks.update_one(
            {"_id": wh["_id"]},
            {"$set": {"lastTriggeredAt": datetime.utcnow(), "lastResponseStatus": 0, "updatedAt": datetime.utcnow()}}
        )

def event_helper(event) -> dict:
    return {
        "id": str(event["_id"]),
        "title": event["title"],
        "description": event.get("description"),
        "date": event["date"],
        "type": event["type"],
        "contactId": event.get("contactId"),
        "organizationId": event.get("organizationId"),
        "isRecurring": event.get("isRecurring", False),
        "recurringPattern": event.get("recurringPattern"),
        "isActive": event.get("isActive", True),
        "createdAt": event.get("createdAt"),
        "updatedAt": event.get("updatedAt")
    }

# Contact Routes
@api_router.post("/contacts/sync", tags=["Contacts"], summary="Sync contacts from device")
async def sync_contacts(contacts: List[Contact]):
    """Import contacts from the device. Duplicates are skipped based on phone number."""
    synced_count = 0
    for contact in contacts:
        # Check if contact already exists by phone
        existing = await db.contacts.find_one({"phone": contact.phone})
        if not existing:
            contact_dict = contact.dict(exclude={"id"})
            contact_dict["createdAt"] = datetime.utcnow()
            contact_dict["updatedAt"] = datetime.utcnow()
            result = await db.contacts.insert_one(contact_dict)
            synced_count += 1
            await dispatch_webhook_event("contact.created", {"id": str(result.inserted_id), "name": contact.name, "phone": contact.phone, "tags": contact.tags})
    return {"message": f"Synced {synced_count} new contacts", "count": synced_count}

@api_router.get("/contacts", tags=["Contacts"], summary="List all contacts with filters")
async def get_contacts(
    search: Optional[str] = None,
    tag: Optional[str] = None,
    favorite: Optional[bool] = None,
    groupId: Optional[str] = None,
    organizationId: Optional[str] = None,
    createdAfter: Optional[str] = None,
    createdBefore: Optional[str] = None
):
    """Get all contacts with optional filters"""
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"notes": {"$regex": search, "$options": "i"}}
        ]
    if tag:
        query["tags"] = tag
    if favorite is not None:
        query["isFavorite"] = favorite
    if groupId:
        group = await db.groups.find_one({"_id": ObjectId(groupId)})
        if group and group.get("contactIds"):
            query["_id"] = {"$in": [ObjectId(cid) for cid in group["contactIds"]]}
    if organizationId:
        query["organizationId"] = organizationId
    if createdAfter:
        query.setdefault("createdAt", {})
        query["createdAt"]["$gte"] = datetime.fromisoformat(createdAfter)
    if createdBefore:
        query.setdefault("createdAt", {})
        query["createdAt"]["$lte"] = datetime.fromisoformat(createdBefore)
    
    contacts = await db.contacts.find(query).sort("name", 1).to_list(1000)
    return [contact_helper(contact) for contact in contacts]

@api_router.get("/contacts/{contact_id}", tags=["Contacts"], summary="Get a single contact")
async def get_contact(contact_id: str):
    """Returns detailed information about a specific contact by ID."""
    contact = await db.contacts.find_one({"_id": ObjectId(contact_id)})
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    return contact_helper(contact)

@api_router.put("/contacts/{contact_id}", tags=["Contacts"], summary="Update a contact")
async def update_contact(contact_id: str, update: ContactUpdate):
    """Partially update a contact. Only provided fields will be modified."""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if update_data:
        update_data["updatedAt"] = datetime.utcnow()
        result = await db.contacts.update_one(
            {"_id": ObjectId(contact_id)},
            {"$set": update_data}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Contact not found")
    
    contact = await db.contacts.find_one({"_id": ObjectId(contact_id)})
    await dispatch_webhook_event("contact.updated", {"id": contact_id, "name": contact.get("name"), "phone": contact.get("phone"), "changes": list(update_data.keys())})
    return contact_helper(contact)

@api_router.delete("/contacts/{contact_id}", tags=["Contacts"], summary="Delete a contact")
async def delete_contact(contact_id: str):
    """Permanently remove a contact from the database."""
    result = await db.contacts.delete_one({"_id": ObjectId(contact_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact not found")
    await dispatch_webhook_event("contact.deleted", {"id": contact_id})
    return {"message": "Contact deleted successfully"}

@api_router.get("/tags", tags=["Contacts"], summary="Get all unique tags")
async def get_all_tags(organizationId: Optional[str] = None):
    """Returns a sorted list of all unique tags used across contacts."""
    query = {}
    if organizationId:
        query["organizationId"] = organizationId
    tags = await db.contacts.distinct("tags", query)
    return {"tags": sorted([tag for tag in tags if tag])}

# Group Routes
@api_router.post("/groups", tags=["Groups"], summary="Create a new group")
async def create_group(group: Group):
    """Create a new contact group with optional color and member list."""
    group_dict = group.dict(exclude={"id"})
    group_dict["createdAt"] = datetime.utcnow()
    result = await db.groups.insert_one(group_dict)
    group_dict["_id"] = result.inserted_id
    return group_helper(group_dict)

@api_router.get("/groups", tags=["Groups"], summary="List all groups")
async def get_groups(organizationId: Optional[str] = None):
    """Returns all groups, optionally filtered by organization."""
    query = {}
    if organizationId:
        query["organizationId"] = organizationId
    groups = await db.groups.find(query).sort("name", 1).to_list(100)
    return [group_helper(group) for group in groups]

@api_router.get("/groups/{group_id}", tags=["Groups"], summary="Get group details with contacts")
async def get_group(group_id: str):
    """Returns a group with its associated contacts populated."""
    group = await db.groups.find_one({"_id": ObjectId(group_id)})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    contact_ids = [ObjectId(cid) for cid in group.get("contactIds", [])]
    contacts = await db.contacts.find({"_id": {"$in": contact_ids}}).to_list(1000)
    
    result = group_helper(group)
    result["contacts"] = [contact_helper(contact) for contact in contacts]
    return result

@api_router.put("/groups/{group_id}", tags=["Groups"], summary="Update a group")
async def update_group(group_id: str, update: GroupUpdate):
    """Partially update group name, color, or member list."""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if update_data:
        result = await db.groups.update_one(
            {"_id": ObjectId(group_id)},
            {"$set": update_data}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Group not found")
    
    group = await db.groups.find_one({"_id": ObjectId(group_id)})
    return group_helper(group)

@api_router.delete("/groups/{group_id}", tags=["Groups"], summary="Delete a group")
async def delete_group(group_id: str):
    """Permanently remove a group. Contacts are not deleted."""
    result = await db.groups.delete_one({"_id": ObjectId(group_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Group not found")
    return {"message": "Group deleted successfully"}

class ImportGroup(BaseModel):
    name: str
    color: Optional[str] = "#4A90E2"
    contactPhones: List[str] = []

@api_router.post("/groups/import", tags=["Groups"], summary="Import groups in bulk")
async def import_groups(groups: List[ImportGroup]):
    """Import multiple groups at once, avoiding duplicates by name. Contacts are resolved by phone number."""
    imported = 0
    skipped = 0
    for group in groups:
        existing = await db.groups.find_one({"name": group.name})
        if existing:
            skipped += 1
            continue
        # Resolve contacts by phone
        contact_ids = []
        for phone in group.contactPhones:
            contact = await db.contacts.find_one({"phone": phone})
            if contact:
                contact_ids.append(str(contact["_id"]))
        
        group_dict = {
            "name": group.name,
            "color": group.color or "#4A90E2",
            "contactIds": contact_ids,
            "createdAt": datetime.utcnow()
        }
        await db.groups.insert_one(group_dict)
        imported += 1
    
    return {"imported": imported, "skipped": skipped, "total": len(groups)}

# Organization Routes
@api_router.post("/organizations", tags=["Organizations"], dependencies=[Depends(require_role("admin"))])
async def create_organization(org: Organization):
    """Create a new organization (admin only). The organization is used to isolate contacts, groups, events, and messages."""
    org_dict = org.dict(exclude={"id"})
    now = datetime.utcnow()
    org_dict["createdAt"] = now
    org_dict["updatedAt"] = now
    result = await db.organizations.insert_one(org_dict)
    org_dict["_id"] = result.inserted_id
    return organization_helper(org_dict)

@api_router.get("/organizations", tags=["Organizations"], dependencies=[Depends(require_role("admin"))])
async def get_organizations():
    """List all organizations (admin only)."""
    orgs = await db.organizations.find().sort("name", 1).to_list(100)
    return [organization_helper(org) for org in orgs]

@api_router.get("/organizations/{org_id}", tags=["Organizations"], dependencies=[Depends(require_role("admin"))])
async def get_organization(org_id: str):
    """Get details of a single organization (admin only)."""
    org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    return organization_helper(org)

@api_router.put("/organizations/{org_id}", tags=["Organizations"], dependencies=[Depends(require_role("admin"))])
async def update_organization(org_id: str, update: OrganizationUpdate):
    """Update an organization's name, color, or description (admin only)."""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if update_data:
        update_data["updatedAt"] = datetime.utcnow()
        result = await db.organizations.update_one(
            {"_id": ObjectId(org_id)},
            {"$set": update_data}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Organization not found")
    
    org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    return organization_helper(org)

@api_router.delete("/organizations/{org_id}", tags=["Organizations"], dependencies=[Depends(require_role("admin"))])
async def delete_organization(org_id: str):
    """Delete an organization (admin only). This does NOT cascade-delete associated data."""
    result = await db.organizations.delete_one({"_id": ObjectId(org_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Organization not found")
    return {"message": "Organization deleted successfully"}

# API Key Routes
@api_router.post("/api-keys", response_model=ApiKeyResponse, tags=["API Keys"], dependencies=[Depends(require_role("admin"))])
async def create_api_key(key_data: ApiKeyCreate, user: dict = Depends(get_current_user)):
    """Create a new API key for external integrations. The full key is only shown once on creation."""
    raw_key = generate_api_key()
    doc = {
        "name": key_data.name,
        "keyHash": hash_api_key(raw_key),
        "scopes": key_data.scopes,
        "userId": user["id"],
        "isActive": True,
        "lastUsedAt": None,
        "createdAt": datetime.utcnow()
    }
    result = await db.api_keys.insert_one(doc)
    return ApiKeyResponse(
        id=str(result.inserted_id),
        name=doc["name"],
        key=raw_key,
        scopes=doc["scopes"],
        isActive=True,
        createdAt=doc["createdAt"]
    )

@api_router.get("/api-keys", tags=["API Keys"], dependencies=[Depends(require_role("admin"))])
async def list_api_keys(user: dict = Depends(get_current_user)):
    """List all API keys for the current user. The actual key values are never returned."""
    docs = await db.api_keys.find({"userId": user["id"]}).sort("createdAt", -1).to_list(100)
    return [api_key_helper(d) for d in docs]

@api_router.delete("/api-keys/{key_id}", tags=["API Keys"], dependencies=[Depends(require_role("admin"))])
async def delete_api_key(key_id: str, user: dict = Depends(get_current_user)):
    """Permanently delete an API key. Integrations using this key will stop working."""
    result = await db.api_keys.delete_one({"_id": ObjectId(key_id), "userId": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="API key not found")
    return {"message": "API key deleted successfully"}

@api_router.post("/api-keys/{key_id}/toggle", tags=["API Keys"], dependencies=[Depends(require_role("admin"))])
async def toggle_api_key(key_id: str, user: dict = Depends(get_current_user)):
    """Enable or disable an API key without deleting it."""
    doc = await db.api_keys.find_one({"_id": ObjectId(key_id), "userId": user["id"]})
    if not doc:
        raise HTTPException(status_code=404, detail="API key not found")
    new_status = not doc.get("isActive", True)
    await db.api_keys.update_one({"_id": ObjectId(key_id)}, {"$set": {"isActive": new_status}})
    return {"id": key_id, "isActive": new_status}

# ─── CRM Integration ────────────────────────────────────
CRM_PROVIDERS = {
    "hubspot": {
        "name": "HubSpot",
        "description": "HubSpot CRM - bidirectional contact sync",
        "default_url": "https://api.hubapi.com",
    },
    "salesforce": {
        "name": "Salesforce",
        "description": "Salesforce CRM - bidirectional contact sync",
        "default_url": "https://your-instance.salesforce.com",
    },
}

@api_router.get("/crm/providers", tags=["CRM"], summary="List available CRM providers")
async def list_crm_providers():
    """Returns a list of supported CRM providers (HubSpot, Salesforce) with metadata."""
    return [
        CrmProviderInfo(id=pid, name=p["name"], description=p["description"])
        for pid, p in CRM_PROVIDERS.items()
    ]

@api_router.post("/crm/integrations", tags=["CRM"], summary="Create a CRM integration")
async def create_crm_integration(integration: CrmIntegrationCreate, user: dict = Depends(get_current_user_or_api_key)):
    """Configure a new CRM integration with API credentials."""
    if integration.provider not in CRM_PROVIDERS:
        raise HTTPException(status_code=400, detail=f"Unsupported CRM provider. Supported: {', '.join(CRM_PROVIDERS.keys())}")
    
    provider = CRM_PROVIDERS[integration.provider]
    doc = {
        "provider": integration.provider,
        "name": integration.name or provider["name"],
        "apiKey": integration.apiKey,
        "apiUrl": integration.apiUrl or provider["default_url"],
        "organizationId": integration.organizationId,
        "userId": user["id"],
        "isActive": True,
        "lastSyncAt": None,
        "lastSyncStatus": None,
        "createdAt": datetime.utcnow(),
        "updatedAt": datetime.utcnow()
    }
    result = await db.crm_integrations.insert_one(doc)
    doc["_id"] = result.inserted_id
    return crm_integration_helper(doc)

@api_router.get("/crm/integrations", tags=["CRM"], summary="List CRM integrations")
async def list_crm_integrations(user: dict = Depends(get_current_user_or_api_key)):
    """Returns all configured CRM integrations for the current user."""
    docs = await db.crm_integrations.find({"userId": user["id"]}).sort("createdAt", -1).to_list(50)
    return [crm_integration_helper(d) for d in docs]

@api_router.put("/crm/integrations/{integration_id}", tags=["CRM"], summary="Update a CRM integration")
async def update_crm_integration(integration_id: str, update: CrmIntegrationUpdate, user: dict = Depends(get_current_user_or_api_key)):
    """Update CRM integration credentials or settings."""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    update_data["updatedAt"] = datetime.utcnow()
    result = await db.crm_integrations.update_one(
        {"_id": ObjectId(integration_id), "userId": user["id"]},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="CRM integration not found")
    doc = await db.crm_integrations.find_one({"_id": ObjectId(integration_id)})
    return crm_integration_helper(doc)

@api_router.delete("/crm/integrations/{integration_id}", tags=["CRM"], summary="Delete a CRM integration")
async def delete_crm_integration(integration_id: str, user: dict = Depends(get_current_user_or_api_key)):
    """Remove a CRM integration configuration."""
    result = await db.crm_integrations.delete_one({"_id": ObjectId(integration_id), "userId": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="CRM integration not found")
    return {"message": "CRM integration deleted successfully"}

async def hubspot_sync_contacts(api_key: str, api_url: str, organization_id: Optional[str] = None) -> dict:
    """Sync contacts with HubSpot CRM"""
    import httpx
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    synced = 0
    errors = 0
    
    async with httpx.AsyncClient(timeout=30) as client:
        # Pull contacts from HubSpot (paginated)
        after = None
        hubspot_contacts = []
        for _ in range(5):  # Max 5 pages
            params: dict = {"limit": 100}
            if after:
                params["after"] = after
            try:
                resp = await client.get(f"{api_url}/crm/v3/objects/contacts", headers=headers, params=params)
                if resp.status_code != 200:
                    errors += 1
                    break
                data = resp.json()
                for result in data.get("results", []):
                    props = result.get("properties", {})
                    hubspot_contacts.append({
                        "id": result["id"],
                        "name": f"{props.get('firstname', '')} {props.get('lastname', '')}".strip(),
                        "phone": props.get("phone", ""),
                        "email": props.get("email", ""),
                    })
                after = data.get("paging", {}).get("next", {}).get("after")
                if not after:
                    break
            except Exception:
                errors += 1
                break
        
        # Push local contacts to HubSpot
        query = {}
        if organization_id:
            query["organizationId"] = organization_id
        local_contacts = await db.contacts.find(query).to_list(500)
        
        for contact in local_contacts:
            try:
                # Check if contact exists in HubSpot by phone
                search_resp = await client.post(
                    f"{api_url}/crm/v3/objects/contacts/search",
                    headers=headers,
                    json={
                        "filterGroups": [{
                            "filters": [{"propertyName": "phone", "operator": "EQ", "value": contact["phone"]}]
                        }]
                    }
                )
                exists = search_resp.status_code == 200 and len(search_resp.json().get("results", [])) > 0
                
                properties = {
                    "firstname": contact["name"].split(" ")[0] if contact["name"] else "",
                    "lastname": " ".join(contact["name"].split(" ")[1:]) if " " in contact["name"] else "",
                    "phone": contact["phone"],
                }
                
                if exists:
                    hubspot_id = search_resp.json()["results"][0]["id"]
                    await client.patch(
                        f"{api_url}/crm/v3/objects/contacts/{hubspot_id}",
                        headers=headers,
                        json={"properties": properties}
                    )
                else:
                    await client.post(
                        f"{api_url}/crm/v3/objects/contacts",
                        headers=headers,
                        json={"properties": properties}
                    )
                synced += 1
            except Exception:
                errors += 1
    
    return {"synced": synced, "errors": errors, "pulled": len(hubspot_contacts)}

async def salesforce_sync_contacts(api_key: str, api_url: str, organization_id: Optional[str] = None) -> dict:
    """Sync contacts with Salesforce CRM"""
    import httpx
    # Salesforce uses OAuth2 - the api_key here is the session token
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    synced = 0
    errors = 0
    
    async with httpx.AsyncClient(timeout=30) as client:
        query = {}
        if organization_id:
            query["organizationId"] = organization_id
        local_contacts = await db.contacts.find(query).to_list(500)
        
        for contact in local_contacts:
            try:
                # Query Salesforce for existing contact by phone
                phone_clean = contact["phone"].replace(" ", "")
                search_resp = await client.get(
                    f"{api_url}/services/data/v58.0/query",
                    headers=headers,
                    params={"q": f"SELECT Id, Name, Phone FROM Contact WHERE Phone = '{phone_clean}'"}
                )
                exists = search_resp.status_code == 200 and len(search_resp.json().get("records", [])) > 0
                
                contact_data = {
                    "LastName": contact["name"] if contact["name"] else "Unknown",
                    "Phone": contact["phone"],
                }
                
                if exists:
                    sf_id = search_resp.json()["records"][0]["Id"]
                    await client.patch(
                        f"{api_url}/services/data/v58.0/sobjects/Contact/{sf_id}",
                        headers=headers,
                        json=contact_data
                    )
                else:
                    await client.post(
                        f"{api_url}/services/data/v58.0/sobjects/Contact",
                        headers=headers,
                        json=contact_data
                    )
                synced += 1
            except Exception:
                errors += 1
    
    return {"synced": synced, "errors": errors, "pulled": 0}

@api_router.post("/crm/integrations/{integration_id}/sync", tags=["CRM"], summary="Sync contacts with CRM")
async def sync_crm_contacts(integration_id: str, user: dict = Depends(get_current_user_or_api_key)):
    """Trigger a bidirectional contact sync with the configured CRM."""
    doc = await db.crm_integrations.find_one({"_id": ObjectId(integration_id), "userId": user["id"]})
    if not doc:
        raise HTTPException(status_code=404, detail="CRM integration not found")
    if not doc.get("isActive"):
        raise HTTPException(status_code=400, detail="CRM integration is disabled")
    
    provider = doc["provider"]
    api_key = doc.get("apiKey", "")
    api_url = doc.get("apiUrl", CRM_PROVIDERS.get(provider, {}).get("default_url", ""))
    org_id = doc.get("organizationId")
    
    try:
        if provider == "hubspot":
            result = await hubspot_sync_contacts(api_key, api_url, org_id)
        elif provider == "salesforce":
            result = await salesforce_sync_contacts(api_key, api_url, org_id)
        else:
            raise HTTPException(status_code=400, detail=f"Unknown provider: {provider}")
        
        status = "success" if result["errors"] == 0 else "partial"
        await db.crm_integrations.update_one(
            {"_id": ObjectId(integration_id)},
            {"$set": {"lastSyncAt": datetime.utcnow(), "lastSyncStatus": status, "updatedAt": datetime.utcnow()}}
        )
        return {"provider": provider, "status": status, **result}
    except Exception as e:
        await db.crm_integrations.update_one(
            {"_id": ObjectId(integration_id)},
            {"$set": {"lastSyncAt": datetime.utcnow(), "lastSyncStatus": "failed", "updatedAt": datetime.utcnow()}}
        )
        raise HTTPException(status_code=500, detail=f"CRM sync failed: {str(e)}")

# ─── Webhook Routes ─────────────────────────────────────
VALID_WEBHOOK_EVENTS = [
    "contact.created", "contact.updated", "contact.deleted",
    "message.scheduled", "message.sent", "message.failed",
    "event.created", "event.upcoming",
]

@api_router.get("/webhooks/events", tags=["Webhooks"], summary="List valid webhook event types")
async def list_webhook_events():
    """Returns the list of event types that can trigger webhooks."""
    return {"events": VALID_WEBHOOK_EVENTS}

@api_router.post("/webhooks", tags=["Webhooks"], summary="Create a webhook")
async def create_webhook(wh: WebhookCreate, user: dict = Depends(get_current_user_or_api_key)):
    """Register a new webhook URL to receive event notifications."""
    for event in wh.events:
        if event not in VALID_WEBHOOK_EVENTS:
            raise HTTPException(status_code=400, detail=f"Invalid event: {event}. Valid: {', '.join(VALID_WEBHOOK_EVENTS)}")
    doc = {
        "url": wh.url,
        "events": wh.events,
        "name": wh.name,
        "secret": wh.secret,
        "userId": user["id"],
        "isActive": True,
        "lastTriggeredAt": None,
        "lastResponseStatus": None,
        "createdAt": datetime.utcnow(),
        "updatedAt": datetime.utcnow()
    }
    result = await db.webhooks.insert_one(doc)
    doc["_id"] = result.inserted_id
    return webhook_helper(doc)

@api_router.get("/webhooks", tags=["Webhooks"], summary="List webhooks")
async def list_webhooks(user: dict = Depends(get_current_user_or_api_key)):
    """Returns all registered webhooks for the current user."""
    docs = await db.webhooks.find({"userId": user["id"]}).sort("createdAt", -1).to_list(50)
    return [webhook_helper(d) for d in docs]

@api_router.put("/webhooks/{webhook_id}", tags=["Webhooks"], summary="Update a webhook")
async def update_webhook(webhook_id: str, update: WebhookUpdate, user: dict = Depends(get_current_user_or_api_key)):
    """Update webhook URL, events, name, or active status."""
    update_data = {k: v for k, v in update.dict(exclude_none=True).items()}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    if "events" in update_data:
        for event in update_data["events"]:
            if event not in VALID_WEBHOOK_EVENTS:
                raise HTTPException(status_code=400, detail=f"Invalid event: {event}")
    update_data["updatedAt"] = datetime.utcnow()
    result = await db.webhooks.update_one(
        {"_id": ObjectId(webhook_id), "userId": user["id"]},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Webhook not found")
    doc = await db.webhooks.find_one({"_id": ObjectId(webhook_id)})
    return webhook_helper(doc)

@api_router.delete("/webhooks/{webhook_id}", tags=["Webhooks"], summary="Delete a webhook")
async def delete_webhook(webhook_id: str, user: dict = Depends(get_current_user_or_api_key)):
    """Remove a webhook."""
    result = await db.webhooks.delete_one({"_id": ObjectId(webhook_id), "userId": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Webhook not found")
    return {"message": "Webhook deleted successfully"}

@api_router.post("/scheduled-messages", tags=["Scheduled Messages"], summary="Create a scheduled message")
async def create_scheduled_message(scheduled_message: ScheduledMessage):
    """Schedule a message to be sent to a group at a specific time. Supports recurring patterns (daily, weekly, monthly)."""
    scheduled_message_dict = scheduled_message.dict(exclude={"id"})
    scheduled_message_dict["createdAt"] = datetime.utcnow()
    scheduled_message_dict["updatedAt"] = datetime.utcnow()
    result = await db.scheduled_messages.insert_one(scheduled_message_dict)
    scheduled_message_dict["_id"] = result.inserted_id
    await dispatch_webhook_event("message.scheduled", {"id": str(result.inserted_id), "groupId": scheduled_message.groupId, "scheduledTime": scheduled_message.scheduledTime.isoformat(), "status": "pending"})
    return scheduled_message_helper(scheduled_message_dict)

@api_router.get("/scheduled-messages", tags=["Scheduled Messages"], summary="List scheduled messages")
async def get_scheduled_messages(group_id: Optional[str] = None, status: Optional[str] = None, active: Optional[bool] = None, organizationId: Optional[str] = None):
    """Returns scheduled messages with optional filters by group, status, and activity."""
    query = {}
    if group_id:
        query["groupId"] = group_id
    if status:
        query["status"] = status
    if active is not None:
        query["isActive"] = active
    if organizationId:
        query["organizationId"] = organizationId
    
    scheduled_messages = await db.scheduled_messages.find(query).sort("scheduledTime", 1).to_list(1000)
    return [scheduled_message_helper(msg) for msg in scheduled_messages]

@api_router.get("/scheduled-messages/{message_id}", tags=["Scheduled Messages"], summary="Get a scheduled message")
async def get_scheduled_message(message_id: str):
    """Returns details of a specific scheduled message."""
    scheduled_message = await db.scheduled_messages.find_one({"_id": ObjectId(message_id)})
    if not scheduled_message:
        raise HTTPException(status_code=404, detail="Scheduled message not found")
    return scheduled_message_helper(scheduled_message)

@api_router.put("/scheduled-messages/{message_id}", tags=["Scheduled Messages"], summary="Update a scheduled message")
async def update_scheduled_message(message_id: str, update: ScheduledMessageUpdate):
    """Partially update a scheduled message (message text, time, recurrence, status)."""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if update_data:
        update_data["updatedAt"] = datetime.utcnow()
        result = await db.scheduled_messages.update_one(
            {"_id": ObjectId(message_id)},
            {"$set": update_data}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Scheduled message not found")
    
    scheduled_message = await db.scheduled_messages.find_one({"_id": ObjectId(message_id)})
    return scheduled_message_helper(scheduled_message)

@api_router.delete("/scheduled-messages/{message_id}", tags=["Scheduled Messages"], summary="Delete a scheduled message")
async def delete_scheduled_message(message_id: str):
    """Permanently remove a scheduled message."""
    result = await db.scheduled_messages.delete_one({"_id": ObjectId(message_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Scheduled message not found")
    return {"message": "Scheduled message deleted successfully"}

@app.post("/api/scheduled-messages/send", tags=["Scheduled Messages"], summary="Send a scheduled message immediately")
async def send_scheduled_message(message_id: str):
    """Trigger immediate delivery of a pending scheduled message."""
    scheduled_message = await db.scheduled_messages.find_one({"_id": ObjectId(message_id)})
    if not scheduled_message:
        raise HTTPException(status_code=404, detail="Scheduled message not found")
    
    try:
        # Get group details
        group = await db.groups.find_one({"_id": ObjectId(scheduled_message["groupId"])})
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        
        # Send message via WhatsApp Business API
        results = await send_whatsapp_message(group, scheduled_message["message"])
        all_ok = all(r.get("success") for r in results) if results else False
        
        new_status = "sent" if all_ok else "failed"
        await db.scheduled_messages.update_one(
            {"_id": ObjectId(message_id)},
            {"$set": {"status": new_status, "updatedAt": datetime.utcnow()}}
        )
        
        await dispatch_webhook_event(f"message.{new_status}", {"id": message_id, "groupId": scheduled_message.get("groupId"), "status": new_status})
        
        if all_ok:
            return {"message": "Message sent successfully", "results": results}
        else:
            raise HTTPException(status_code=500, detail=f"Failed to send to some contacts: {results}")
    except Exception as e:
        await db.scheduled_messages.update_one(
            {"_id": ObjectId(message_id)},
            {"$set": {"status": "failed", "updatedAt": datetime.utcnow()}}
        )
        await dispatch_webhook_event("message.failed", {"id": message_id, "error": str(e)})
        raise HTTPException(status_code=500, detail=str(e))

def whatsapp_config_helper(doc) -> dict:
    return {
        "id": str(doc["_id"]),
        "phoneNumberId": doc["phoneNumberId"],
        "businessAccountId": doc.get("businessAccountId"),
        "isActive": doc.get("isActive", True),
        "organizationId": doc.get("organizationId"),
        "createdAt": doc["createdAt"],
        "updatedAt": doc["updatedAt"]
    }

def whatsapp_message_status_helper(doc) -> dict:
    return {
        "id": str(doc["_id"]),
        "externalMessageId": doc["externalMessageId"],
        "scheduledMessageId": doc.get("scheduledMessageId"),
        "recipientPhone": doc["recipientPhone"],
        "status": doc["status"],
        "timestamp": doc["timestamp"]
    }

WHATSAPP_API_BASE = "https://graph.facebook.com/v18.0"

async def _get_whatsapp_config(organization_id: Optional[str] = None) -> Optional[dict]:
    """Get active WhatsApp Business API config"""
    query = {"isActive": True}
    if organization_id:
        query["organizationId"] = organization_id
    return await db.whatsapp_config.find_one(query)

async def send_single_whatsapp_message(phone: str, message: str, scheduled_message_id: Optional[str] = None, organization_id: Optional[str] = None) -> dict:
    """Send a WhatsApp message via Cloud API and track its status"""
    config = await _get_whatsapp_config(organization_id)
    if not config:
        return {"success": False, "error": "WhatsApp not configured"}
    
    import httpx
    url = f"{WHATSAPP_API_BASE}/{config['phoneNumberId']}/messages"
    headers = {
        "Authorization": f"Bearer {config['accessToken']}",
        "Content-Type": "application/json"
    }
    body = {
        "messaging_product": "whatsapp",
        "to": phone,
        "type": "text",
        "text": {"body": message}
    }
    
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.post(url, json=body, headers=headers)
            data = resp.json()
            if resp.status_code == 200 or resp.status_code == 201:
                wa_id = data.get("messages", [{}])[0].get("id", "")
                if wa_id:
                    await db.whatsapp_message_status.insert_one({
                        "externalMessageId": wa_id,
                        "scheduledMessageId": scheduled_message_id,
                        "recipientPhone": phone,
                        "status": "sent",
                        "timestamp": datetime.utcnow(),
                        "organizationId": organization_id
                    })
                return {"success": True, "waMessageId": wa_id}
            else:
                error = data.get("error", {}).get("message", str(data))
                return {"success": False, "error": error}
    except Exception as e:
        return {"success": False, "error": str(e)}

async def send_whatsapp_message(group: dict, message: str):
    """Send WhatsApp message to all contacts in a group via Cloud API"""
    contact_ids = [ObjectId(cid) for cid in group.get("contactIds", [])]
    if not contact_ids:
        return True
    
    contacts = await db.contacts.find({"_id": {"$in": contact_ids}}).to_list(1000)
    org_id = group.get("organizationId")
    results = []
    
    for contact in contacts:
        phone = contact.get("phone", "")
        if phone:
            result = await send_single_whatsapp_message(phone, message, organization_id=org_id)
            results.append(result)
    
    return all(r.get("success") for r in results)

# WhatsApp Config Routes
@api_router.post("/whatsapp/config", tags=["WhatsApp"], summary="Configure WhatsApp Business API")
async def create_whatsapp_config(config: WhatsAppConfigCreate, user: dict = Depends(get_current_user_or_api_key)):
    """Save WhatsApp Cloud API credentials. Only one active config per organization."""
    # Deactivate any existing config for this org
    existing_query = {}
    if config.organizationId:
        existing_query["organizationId"] = config.organizationId
    await db.whatsapp_config.update_many(existing_query, {"$set": {"isActive": False}})
    
    doc = {
        "phoneNumberId": config.phoneNumberId,
        "accessToken": config.accessToken,
        "businessAccountId": config.businessAccountId,
        "webhookSecret": config.webhookSecret,
        "organizationId": config.organizationId,
        "userId": user["id"],
        "isActive": True,
        "createdAt": datetime.utcnow(),
        "updatedAt": datetime.utcnow()
    }
    result = await db.whatsapp_config.insert_one(doc)
    doc["_id"] = result.inserted_id
    return whatsapp_config_helper(doc)

@api_router.get("/whatsapp/config", tags=["WhatsApp"], summary="Get WhatsApp configuration")
async def get_whatsapp_config(organizationId: Optional[str] = None, user: dict = Depends(get_current_user_or_api_key)):
    """Returns the active WhatsApp Business API configuration."""
    query = {"userId": user["id"], "isActive": True}
    if organizationId:
        query["organizationId"] = organizationId
    doc = await db.whatsapp_config.find_one(query)
    if not doc:
        raise HTTPException(status_code=404, detail="WhatsApp not configured")
    return whatsapp_config_helper(doc)

@api_router.put("/whatsapp/config", tags=["WhatsApp"], summary="Update WhatsApp configuration")
async def update_whatsapp_config(update: WhatsAppConfigUpdate, user: dict = Depends(get_current_user_or_api_key)):
    """Update existing WhatsApp configuration."""
    update_data = {k: v for k, v in update.dict(exclude_none=True).items()}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    update_data["updatedAt"] = datetime.utcnow()
    result = await db.whatsapp_config.update_one(
        {"userId": user["id"], "isActive": True},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="WhatsApp config not found")
    doc = await db.whatsapp_config.find_one({"userId": user["id"], "isActive": True})
    return whatsapp_config_helper(doc)

@api_router.get("/whatsapp/status", tags=["WhatsApp"], summary="Get message delivery status")
async def get_whatsapp_message_status(phone: Optional[str] = None, limit: int = 50, user: dict = Depends(get_current_user_or_api_key)):
    """Returns delivery status of WhatsApp messages sent through the API."""
    query = {}
    if phone:
        query["recipientPhone"] = phone
    docs = await db.whatsapp_message_status.find(query).sort("timestamp", -1).to_list(limit)
    return [whatsapp_message_status_helper(d) for d in docs]

# WhatsApp Webhook receiver (outside /api prefix for Meta verification)
@app.get("/whatsapp/webhook", tags=["WhatsApp"])
async def whatsapp_webhook_verify(
    hub_mode: Optional[str] = None,
    hub_verify_token: Optional[str] = None,
    hub_challenge: Optional[str] = None
):
    """WhatsApp Cloud API webhook verification endpoint (GET)"""
    config = await _get_whatsapp_config()
    if not config:
        raise HTTPException(status_code=403, detail="WhatsApp not configured")
    expected_token = config.get("webhookSecret", "")
    if hub_mode == "subscribe" and hub_verify_token == expected_token:
        return int(hub_challenge) if hub_challenge else 200
    raise HTTPException(status_code=403, detail="Verification failed")

@app.post("/whatsapp/webhook", tags=["WhatsApp"])
async def whatsapp_webhook_receive(payload: dict):
    """Receive delivery status updates from WhatsApp Cloud API"""
    try:
        entries = payload.get("entry", [])
        for entry in entries:
            changes = entry.get("changes", [])
            for change in changes:
                value = change.get("value", {})
                statuses = value.get("statuses", [])
                for status in statuses:
                    wa_id = status.get("id", "")
                    status_name = status.get("status", "")  # sent, delivered, read, failed
                    timestamp = status.get("timestamp", "")
                    
                    if wa_id:
                        await db.whatsapp_message_status.update_one(
                            {"externalMessageId": wa_id},
                            {"$set": {
                                "status": status_name,
                                "timestamp": datetime.fromtimestamp(int(timestamp)) if timestamp else datetime.utcnow(),
                                "updatedAt": datetime.utcnow()
                            }}
                        )
                        
                        # Also update scheduled message if linked
                        msg_doc = await db.whatsapp_message_status.find_one({"externalMessageId": wa_id})
                        if msg_doc and msg_doc.get("scheduledMessageId"):
                            sms_status = "sent" if status_name in ("sent", "delivered", "read") else "failed"
                            await db.scheduled_messages.update_one(
                                {"_id": ObjectId(msg_doc["scheduledMessageId"])},
                                {"$set": {"status": sms_status, "updatedAt": datetime.utcnow()}}
                            )
    except Exception:
        pass
    
    return {"status": "ok"}

# Backup and Restore Routes
@api_router.get("/backup", tags=["Backup & Export"], summary="Create a full backup")
async def create_backup():
    """Creates a ZIP backup containing all contacts, groups, and scheduled messages as JSON."""
    try:
        # Create backup data
        backup_data = {
            "backupId": str(ObjectId()),
            "createdAt": datetime.utcnow(),
            "version": "1.0",
            "contacts": [],
            "groups": [],
            "scheduledMessages": []
        }
        
        # Export contacts
        contacts = await db.contacts.find().to_list(1000)
        backup_data["contacts"] = [contact_helper(contact) for contact in contacts]
        
        # Export groups
        groups = await db.groups.find().to_list(1000)
        backup_data["groups"] = [group_helper(group) for group in groups]
        
        # Export scheduled messages
        scheduled_messages = await db.scheduled_messages.find().to_list(1000)
        backup_data["scheduledMessages"] = [scheduled_message_helper(msg) for msg in scheduled_messages]
        
        # Create ZIP file
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add backup data as JSON
            zip_file.writestr('backup.json', json.dumps(backup_data, indent=2, default=str))
            
            # Add contact photos if they exist
            for contact in backup_data["contacts"]:
                if contact.get("photo"):
                    # In a real implementation, you'd handle image data properly
                    pass
        
        zip_buffer.seek(0)
        
        return {
            "backupId": backup_data["backupId"],
            "createdAt": backup_data["createdAt"],
            "size": len(zip_buffer.getvalue()),
            "data": zip_buffer.getvalue()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup failed: {str(e)}")

@api_router.post("/restore", tags=["Backup & Export"], summary="Restore from backup")
async def restore_backup(backup_data: dict):
    """Restore contacts, groups, and scheduled messages from a backup JSON object."""
    try:
        # Clear existing data (optional - add confirmation in production)
        # await db.contacts.delete_many({})
        # await db.groups.delete_many({})
        # await db.scheduled_messages.delete_many({})
        
        # Restore contacts
        if backup_data.get("contacts"):
            for contact in backup_data["contacts"]:
                await db.contacts.insert_one(contact)
        
        # Restore groups
        if backup_data.get("groups"):
            for group in backup_data["groups"]:
                await db.groups.insert_one(group)
        
        # Restore scheduled messages
        if backup_data.get("scheduledMessages"):
            for message in backup_data["scheduledMessages"]:
                await db.scheduled_messages.insert_one(message)
        
        return {
            "message": "Backup restored successfully",
            "contactsCount": len(backup_data.get("contacts", [])),
            "groupsCount": len(backup_data.get("groups", [])),
            "scheduledMessagesCount": len(backup_data.get("scheduledMessages", []))
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")

@api_router.get("/export/contacts", tags=["Backup & Export"], summary="Export contacts as CSV")
async def export_contacts():
    """Download all contacts in CSV format for use in spreadsheets."""
    try:
        contacts = await db.contacts.find().to_list(1000)
        
        # Generate CSV content
        csv_lines = ["Name,Phone,Email,Tags,Favorite,Notes"]
        for contact in contacts:
            tags = ";".join(contact.get("tags", []))
            csv_lines.append(f'"{contact.get("name", "")}","{contact.get("phone", "")}","{tags}","{contact.get("isFavorite", "")}","{contact.get("notes", "")}"')
        
        csv_content = "\n".join(csv_lines)
        
        return {
            "filename": f"contacts_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "content": csv_content,
            "contentType": "text/csv"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@api_router.get("/export/groups", tags=["Backup & Export"], summary="Export groups as JSON")
async def export_groups():
    """Download all groups in JSON format."""
    try:
        groups = await db.groups.find().to_list(1000)
        
        return {
            "filename": f"groups_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "content": json.dumps(groups, indent=2, default=str),
            "contentType": "application/json"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@api_router.get("/export/events", tags=["Backup & Export"], summary="Export events as CSV")
async def export_events():
    """Download all events in CSV format."""
    try:
        events = await db.events.find().sort("date", 1).to_list(1000)
        csv_lines = ["Title,Type,Date,ContactId,Description,isActive"]
        for ev in events:
            date_str = ev.get("date", "")
            if hasattr(date_str, "isoformat"):
                date_str = date_str.isoformat()
            csv_lines.append(
                f'"{ev.get("title", "")}","{ev.get("type", "")}","{date_str}",'
                f'"{ev.get("contactId", "")}","{ev.get("description", "")}","{ev.get("isActive", True)}"'
            )
        csv_content = "\n".join(csv_lines)
        return {
            "filename": f"events_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "content": csv_content,
            "contentType": "text/csv"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@api_router.get("/export/scheduled-messages", tags=["Backup & Export"], summary="Export scheduled messages as CSV")
async def export_scheduled_messages():
    """Download all scheduled messages in CSV format."""
    try:
        msgs = await db.scheduled_messages.find().sort("scheduledTime", 1).to_list(1000)
        csv_lines = ["Message,ScheduledTime,GroupId,Status,Recurring,Pattern"]
        for msg in msgs:
            time_str = msg.get("scheduledTime", "")
            if hasattr(time_str, "isoformat"):
                time_str = time_str.isoformat()
            csv_lines.append(
                f'"{msg.get("message", "")}","{time_str}","{msg.get("groupId", "")}",'
                f'"{msg.get("status", "pending")}","{msg.get("isRecurring", False)}","{msg.get("recurringPattern", "")}"'
            )
        csv_content = "\n".join(csv_lines)
        return {
            "filename": f"scheduled_messages_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "content": csv_content,
            "contentType": "text/csv"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@api_router.get("/export/all", tags=["Backup & Export"], summary="Export all data as XLSX (Excel)")
async def export_all():
    """Download all data (contacts, groups, events, scheduled messages) in a single Excel file with multiple sheets."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        
        # ── Contacts sheet ──
        ws_contacts = wb.active
        ws_contacts.title = "Contacts"
        ws_contacts.append(["Name", "Phone", "Tags", "Favorite", "Notes", "CreatedAt"])
        contacts = await db.contacts.find().to_list(1000)
        for c in contacts:
            ws_contacts.append([
                c.get("name", ""),
                c.get("phone", ""),
                ";".join(c.get("tags", [])),
                "Yes" if c.get("isFavorite") else "No",
                c.get("notes", ""),
                str(c.get("createdAt", ""))
            ])
        
        # ── Groups sheet ──
        ws_groups = wb.create_sheet("Groups")
        ws_groups.append(["Name", "Color", "MemberCount", "CreatedAt"])
        groups = await db.groups.find().to_list(1000)
        for g in groups:
            ws_groups.append([
                g.get("name", ""),
                g.get("color", ""),
                len(g.get("contactIds", [])),
                str(g.get("createdAt", ""))
            ])
        
        # ── Events sheet ──
        ws_events = wb.create_sheet("Events")
        ws_events.append(["Title", "Type", "Date", "Description", "ContactId", "Recurring", "Active"])
        events = await db.events.find().sort("date", 1).to_list(1000)
        for ev in events:
            date_str = str(ev.get("date", ""))
            ws_events.append([
                ev.get("title", ""),
                ev.get("type", ""),
                date_str,
                ev.get("description", ""),
                ev.get("contactId", ""),
                "Yes" if ev.get("isRecurring") else "No",
                "Yes" if ev.get("isActive", True) else "No"
            ])
        
        # ── Scheduled Messages sheet ──
        ws_sms = wb.create_sheet("ScheduledMessages")
        ws_sms.append(["Message", "ScheduledTime", "GroupId", "Status", "Recurring", "Pattern"])
        sms_list = await db.scheduled_messages.find().sort("scheduledTime", 1).to_list(1000)
        for msg in sms_list:
            ws_sms.append([
                msg.get("message", ""),
                str(msg.get("scheduledTime", "")),
                msg.get("groupId", ""),
                msg.get("status", "pending"),
                "Yes" if msg.get("isRecurring") else "No",
                msg.get("recurringPattern", "")
            ])
        
        xlsx_buffer = BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        return {
            "filename": f"whatsapp_organizer_full_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "content": xlsx_buffer.getvalue(),
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@app.get("/download/backup", tags=["Backup & Export"], summary="Download backup as ZIP")
async def download_backup():
    """Generate and download a complete backup as a ZIP archive."""
    try:
        backup_response = await create_backup()
        
        return {
            "filename": f"whatsapp_organizer_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            "content": backup_response["data"],
            "contentType": "application/zip"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")

# Event Routes
@api_router.post("/events", tags=["Events"], summary="Create a new event")
async def create_event(event: Event):
    """Create a new event (birthday, anniversary, custom) with optional recurrence."""
    event_dict = event.dict(exclude={"id"})
    event_dict["createdAt"] = datetime.utcnow()
    event_dict["updatedAt"] = datetime.utcnow()
    result = await db.events.insert_one(event_dict)
    event_dict["_id"] = result.inserted_id
    await dispatch_webhook_event("event.created", {"id": str(result.inserted_id), "title": event.title, "type": event.type, "date": event.date.isoformat() if event.date else None})
    return event_helper(event_dict)

@api_router.get("/events", tags=["Events"], summary="List all events with filters")
async def get_events(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    type: Optional[str] = None,
    contact_id: Optional[str] = None,
    active: Optional[bool] = None,
    organizationId: Optional[str] = None
):
    """Returns events filtered by date range, type, contact, and active status."""
    query = {}
    
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = start_date
        if end_date:
            query["date"]["$lte"] = end_date
    
    if type:
        query["type"] = type
    
    if contact_id:
        query["contactId"] = contact_id
    
    if active is not None:
        query["isActive"] = active
    
    if organizationId:
        query["organizationId"] = organizationId
    
    events = await db.events.find(query).sort("date", 1).to_list(1000)
    return [event_helper(event) for event in events]

@api_router.get("/events/{event_id}", tags=["Events"], summary="Get a single event")
async def get_event(event_id: str):
    """Returns details of a specific event by ID."""
    event = await db.events.find_one({"_id": ObjectId(event_id)})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    return event_helper(event)

@api_router.put("/events/{event_id}", tags=["Events"], summary="Update an event")
async def update_event(event_id: str, update: EventUpdate):
    """Partially update event details (title, date, type, recurrence, etc.)."""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if update_data:
        update_data["updatedAt"] = datetime.utcnow()
        result = await db.events.update_one(
            {"_id": ObjectId(event_id)},
            {"$set": update_data}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Event not found")
    
    event = await db.events.find_one({"_id": ObjectId(event_id)})
    return event_helper(event)

@api_router.delete("/events/{event_id}", tags=["Events"], summary="Delete an event")
async def delete_event(event_id: str):
    """Permanently remove an event."""
    result = await db.events.delete_one({"_id": ObjectId(event_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted successfully"}

@api_router.get("/events/upcoming", tags=["Events"], summary="Get upcoming events")
async def get_upcoming_events(days_ahead: int = 7, organizationId: Optional[str] = None):
    """Returns active events within the next N days (default 7)."""
    start_date = datetime.utcnow()
    end_date = start_date + timedelta(days=days_ahead)
    
    query = {
        "date": {"$gte": start_date, "$lte": end_date},
        "isActive": True
    }
    if organizationId:
        query["organizationId"] = organizationId
    
    events = await db.events.find(query).sort("date", 1).to_list(100)
    
    return [event_helper(event) for event in events]

@app.post("/api/events/birthday-check", tags=["Events"], summary="Auto-detect birthdays and create events")
async def check_birthdays():
    """Scans contacts with birthday field and creates yearly recurring events if they don't exist."""
    try:
        today = datetime.utcnow()
        
        # Get contacts with birthdays
        contacts = await db.contacts.find({
            "birthday": {"$exists": True}
        }).to_list(1000)
        
        birthday_count = 0
        for contact in contacts:
            if contact.get("birthday"):
                # Parse birthday date
                birthday_date = datetime.fromisoformat(contact["birthday"].replace('Z', '+00:00'))
                
                # Check if birthday event already exists for this year
                existing_event = await db.events.find_one({
                    "contactId": contact["_id"],
                    "type": "birthday",
                    "date": {
                        "$gte": datetime(today.year, birthday_date.month, birthday_date.day, 0, 0, 0),
                        "$lt": datetime(today.year, birthday_date.month, birthday_date.day + 1, 0, 0, 0)
                    }
                })
                
                if not existing_event:
                    # Create birthday event
                    birthday_event = {
                        "title": f"Aniversário de {contact['name']}",
                        "description": f"Hoje é o aniversário de {contact['name']}!",
                        "date": datetime(today.year, birthday_date.month, birthday_date.day, 9, 0, 0),  # 9 AM
                        "type": "birthday",
                        "contactId": str(contact["_id"]),
                        "isRecurring": True,
                        "recurringPattern": "yearly",
                        "isActive": True,
                        "createdAt": datetime.utcnow(),
                        "updatedAt": datetime.utcnow()
                    }
                    
                    result = await db.events.insert_one(birthday_event)
                    birthday_event["_id"] = result.inserted_id
                    birthday_count += 1
                    await dispatch_webhook_event("event.created", {"id": str(result.inserted_id), "title": birthday_event["title"], "type": "birthday", "date": birthday_event["date"].isoformat()})
        
        return {"message": f"Created {birthday_count} birthday events"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Birthday check failed: {str(e)}")

@app.post("/api/events/notification", tags=["Events"], summary="Send a notification for an event")
async def send_event_notification(event_id: str):
    """Trigger a push notification for a specific event. (Placeholder - requires push service integration.)"""
    try:
        event = await db.events.find_one({"_id": ObjectId(event_id)})
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        # Get contact details if associated
        contact_name = "Evento"
        if event.get("contactId"):
            contact = await db.contacts.find_one({"_id": ObjectId(event["contactId"])})
            if contact:
                contact_name = contact["name"]
        
        # Send notification (placeholder - would integrate with push notification service)
        notification_data = {
            "title": event["title"],
            "body": event.get("description", f"Evento de {event['type']} para {contact_name}"),
            "data": {
                "eventId": str(event["_id"]),
                "type": event["type"],
                "contactId": event.get("contactId")
            }
        }
        
        # In a real implementation, this would send push notifications
        print(f"Notification sent: {notification_data}")
        
        return {"message": "Notification sent successfully"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Notification failed: {str(e)}")

# Report Routes
@api_router.get("/reports/contacts-summary", tags=["Reports"], summary="Get contacts summary statistics")
async def get_contacts_summary(organizationId: Optional[str] = None):
    """Returns dashboard metrics: total contacts, favorites, groups, events, pending messages, new this week, and tag breakdown."""
    base_query = {}
    if organizationId:
        base_query["organizationId"] = organizationId
    
    total_contacts = await db.contacts.count_documents(base_query)
    
    favorite_query = {**base_query, "isFavorite": True}
    total_favorites = await db.contacts.count_documents(favorite_query)
    
    # Count by tag
    tag_pipeline = [
        {"$match": base_query},
        {"$unwind": "$tags"},
        {"$group": {"_id": "$tags", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 20}
    ]
    tags_cursor = db.contacts.aggregate(tag_pipeline)
    tags_data = await tags_cursor.to_list(20)
    tags_summary = [{"tag": t["_id"], "count": t["count"]} for t in tags_data]
    
    # Groups count
    total_groups = await db.groups.count_documents(base_query)
    
    # Events count
    total_events = await db.events.count_documents(base_query)
    
    # Scheduled messages count
    sms_query = {**base_query, "status": "pending"}
    pending_messages = await db.scheduled_messages.count_documents(sms_query)
    total_messages = await db.scheduled_messages.count_documents(base_query)
    
    # Contacts created last 7 days
    week_ago = datetime.utcnow() - timedelta(days=7)
    new_this_week = await db.contacts.count_documents({
        **base_query,
        "createdAt": {"$gte": week_ago}
    })
    
    return {
        "totalContacts": total_contacts,
        "totalFavorites": total_favorites,
        "totalGroups": total_groups,
        "totalEvents": total_events,
        "pendingMessages": pending_messages,
        "totalScheduledMessages": total_messages,
        "newContactsThisWeek": new_this_week,
        "tagsBreakdown": tags_summary
    }

@api_router.get("/reports/activity", tags=["Reports"], summary="Get activity over time")
async def get_activity(days: int = 30, organizationId: Optional[str] = None):
    """Returns daily contact and event creation counts for the specified number of days (default 30)."""
    base_query = {}
    if organizationId:
        base_query["organizationId"] = organizationId
    
    since = datetime.utcnow() - timedelta(days=days)
    base_query["createdAt"] = {"$gte": since}
    
    # Group contacts by day
    pipeline = [
        {"$match": base_query},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$createdAt"}},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]
    cursor = db.contacts.aggregate(pipeline)
    daily_data = await cursor.to_list(days)
    
    daily_activity = [{"date": d["_id"], "count": d["count"]} for d in daily_data]
    
    # Also get event activity
    event_query = {**base_query}
    events_pipeline = [
        {"$match": event_query},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$createdAt"}},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]
    event_cursor = db.events.aggregate(events_pipeline)
    event_data = await event_cursor.to_list(days)
    event_activity = [{"date": e["_id"], "count": e["count"]} for e in event_data]
    
    return {
        "daily": daily_activity,
        "events": event_activity,
        "periodDays": days
    }

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(RateLimitMiddleware)

# Include all routes (must be after all @api_router decorators)
app.include_router(api_router)

logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    if 'client' in dir() and client is not None:
        client.close()

# Background scheduler for checking scheduled messages
from apscheduler.schedulers.asyncio import AsyncIOScheduler

scheduler = AsyncIOScheduler()

@app.on_event("startup")
async def start_scheduler():
    scheduler.start()

async def check_and_send_scheduled_messages():
    """Check for scheduled messages that need to be sent"""
    try:
        now = datetime.now(timezone.utc)
        pending_messages = await db.scheduled_messages.find({
            "scheduledTime": {"$lte": now},
            "status": "pending",
            "isActive": True
        }).to_list(100)
        
        for message in pending_messages:
            try:
                # Get group details
                group = await db.groups.find_one({"_id": ObjectId(message["groupId"])})
                if group:
                    # Send message
                    await send_whatsapp_message(group, message["message"])
                    
                    # Update status
                    await db.scheduled_messages.update_one(
                        {"_id": ObjectId(message["_id"])},
                        {"$set": {"status": "sent", "updatedAt": datetime.utcnow()}}
                    )
                    
                    # Handle recurring messages
                    if message.get("isRecurring") and message.get("recurringPattern"):
                        await schedule_next_recurring_message(message)
                
            except Exception as e:
                await db.scheduled_messages.update_one(
                    {"_id": ObjectId(message["_id"])},
                    {"$set": {"status": "failed", "updatedAt": datetime.utcnow()}}
                )
                
    except Exception as e:
        logger.error(f"Error checking scheduled messages: {e}")

async def schedule_next_recurring_message(message):
    """Schedule the next occurrence of a recurring message"""
    try:
        next_time = message["scheduledTime"]
        
        if message.get("recurringPattern") == "daily":
            next_time += timedelta(days=1)
        elif message.get("recurringPattern") == "weekly":
            next_time += timedelta(weeks=1)
        elif message.get("recurringPattern") == "monthly":
            next_time += timedelta(days=30)
        
        # Create new scheduled message
        new_message = {
            "groupId": message["groupId"],
            "message": message["message"],
            "scheduledTime": next_time,
            "isRecurring": True,
            "recurringPattern": message.get("recurringPattern"),
            "isActive": True,
            "status": "pending",
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow()
        }
        
        await db.scheduled_messages.insert_one(new_message)
        
    except Exception as e:
        logger.error(f"Error scheduling next recurring message: {e}")

async def check_and_send_event_notifications():
    """Check for events that need notifications"""
    try:
        now = datetime.now(timezone.utc)
        tomorrow = now + timedelta(days=1)
        
        # Get events happening tomorrow
        events = await db.events.find({
            "date": {"$gte": now, "$lt": tomorrow},
            "isActive": True,
            "type": {"$in": ["birthday", "anniversary"]}
        }).to_list(50)
        
        for event in events:
            try:
                await send_event_notification(str(event["_id"]))
            except Exception as e:
                logger.error(f"Error sending notification for event {event['_id']}: {e}")
                
    except Exception as e:
        logger.error(f"Error checking event notifications: {e}")

# Schedule the checkers to run at different intervals
scheduler.add_job(check_and_send_scheduled_messages, 'interval', minutes=1)
scheduler.add_job(check_and_send_event_notifications, 'interval', hours=1)
scheduler.add_job(check_birthdays, 'cron', hour=0, minute=0)  # Run daily at midnight


